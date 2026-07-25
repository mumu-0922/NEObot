#!/usr/bin/env python3
"""Generate the reversible synthetic source corpus for the RAG promotion draft."""

from __future__ import annotations

import argparse
import functools
import hashlib
import json
import os
import posixpath
import re
import shutil
import subprocess
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

SCHEMA = "neo-chat.rag-evaluation-source-corpus.v1"
GENERATED_AT = "2026-07-24T00:00:00Z"
LANES = ("pdf", "docx", "pptx", "xlsx", "code")
MIMES = {
    "pdf": "application/pdf",
    "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "pptx": "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "code": "text/markdown",
}
EXTENSIONS = {
    "pdf": "pdf",
    "docx": "docx",
    "pptx": "pptx",
    "xlsx": "xlsx",
    "code": "md",
}
TOPICS = {
    "pdf": (
        "青岚微电网运行纪要",
        "北辰冷链巡检报告",
        "赤霄防汛传感器手册",
        "云栖公交调度规程",
        "星河档案馆迁移方案",
        "Atlas Port Safety Review",
        "Beacon Water Plant Handbook",
        "Cedar Laboratory Continuity Plan",
        "Delta Rail Maintenance Report",
        "Ember Library Preservation Guide",
    ),
    "docx": (
        "青岚事件响应制度",
        "北辰采购验收规范",
        "赤霄门诊排班手册",
        "云栖工厂质检办法",
        "星河科研评审细则",
        "Atlas Incident Response Manual",
        "Beacon Procurement Standard",
        "Cedar Clinic Scheduling Guide",
        "Delta Factory Quality Procedure",
        "Ember Grant Review Policy",
    ),
    "pptx": (
        "青岚产品发布简报",
        "北辰博物馆展陈计划",
        "赤霄港口升级路线图",
        "云栖教师培训方案",
        "星河社区能源汇报",
        "Atlas Product Launch Brief",
        "Beacon Museum Exhibit Plan",
        "Cedar Harbor Upgrade Roadmap",
        "Delta Teacher Training Program",
        "Ember Community Energy Review",
    ),
    "xlsx": (
        "青岚备件库存台账",
        "北辰项目预算表",
        "赤霄能耗计量簿",
        "云栖用水核算表",
        "星河车队维护表",
        "Atlas Spare Parts Ledger",
        "Beacon Project Budget",
        "Cedar Energy Meter Register",
        "Delta Water Usage Workbook",
        "Ember Fleet Maintenance Table",
    ),
    "code": (
        "青岚限流服务规范",
        "北辰特性开关契约",
        "赤霄备份调度协议",
        "云栖图像管线配置",
        "星河审计导出接口",
        "Atlas Rate Limiter Specification",
        "Beacon Feature Flag Contract",
        "Cedar Backup Scheduler Protocol",
        "Delta Image Pipeline Configuration",
        "Ember Audit Export API",
    ),
}


def stable_slug(lane: str, index: int, language: str) -> str:
    return f"rag-eval-{lane}-{language}-{index + 1:02d}"


def build_document(lane: str, index: int) -> dict[str, Any]:
    language = "zh" if index < 5 else "en"
    lane_number = LANES.index(lane) + 1
    local_index = index + 1
    doc_id = f"RAGEVAL-{lane.upper()}-{language.upper()}-{local_index:02d}"
    capacity = 120 + lane_number * 37 + local_index * 13
    threshold = 64 + lane_number * 3 + local_index
    sla_minutes = 12 + lane_number * 4 + local_index * 2
    retention_days = 45 + lane_number * 11 + local_index * 7
    day = 2 + lane_number * 3 + local_index
    effective_date = f"2026-08-{day:02d}"
    if language == "zh":
        owners = ("青岚运营组", "北辰质量组", "赤霄保障组", "云栖数据组", "星河审计组")
        locations = ("苏州 A 区", "成都 B 区", "青岛 C 区", "武汉 D 区", "西安 E 区")
        unit = "标准单元"
        owner = owners[index]
        location = locations[index]
        exception_code = f"例外-{lane_number}{local_index:02d}-甲"
        cross_answer = f"容量 {capacity} {unit}，触发阈值 {threshold}%"
        facts = (
            ("F01", "项目名称", TOPICS[lane][index], "overview", ("short_fact",)),
            ("F02", "责任团队", owner, "overview", ("short_fact",)),
            ("F03", "核定容量", f"{capacity} {unit}", "metrics", ("exact_numeric",)),
            ("F04", "触发阈值", f"{threshold}%", "metrics", ("exact_numeric",)),
            ("F05", "生效日期", effective_date, "schedule", ("short_fact",)),
            ("F06", "执行地点", location, "schedule", ("short_fact",)),
            ("F07", "例外代码", exception_code, "exceptions", ("short_fact",)),
            (
                "F08",
                "容量与阈值联查",
                cross_answer,
                "cross_reference",
                ("cross_section", "exact_numeric"),
            ),
            (
                "F09",
                "升级时限",
                f"{sla_minutes} 分钟",
                "exceptions",
                ("exact_numeric",),
            ),
            (
                "F10",
                "记录保留期",
                f"{retention_days} 天",
                "retention",
                ("exact_numeric",),
            ),
        )
    else:
        owners = (
            "Atlas Operations",
            "Beacon Quality",
            "Cedar Reliability",
            "Delta Data",
            "Ember Audit",
        )
        locations = (
            "Austin Zone A",
            "Boston Zone B",
            "Chicago Zone C",
            "Denver Zone D",
            "Eugene Zone E",
        )
        unit = "standard units"
        owner = owners[index - 5]
        location = locations[index - 5]
        exception_code = f"EX-{lane_number}{local_index:02d}-A"
        cross_answer = f"capacity {capacity} {unit} with trigger threshold {threshold}%"
        facts = (
            ("F01", "Project name", TOPICS[lane][index], "overview", ("short_fact",)),
            ("F02", "Owning team", owner, "overview", ("short_fact",)),
            (
                "F03",
                "Approved capacity",
                f"{capacity} {unit}",
                "metrics",
                ("exact_numeric",),
            ),
            (
                "F04",
                "Trigger threshold",
                f"{threshold}%",
                "metrics",
                ("exact_numeric",),
            ),
            ("F05", "Effective date", effective_date, "schedule", ("short_fact",)),
            ("F06", "Operating location", location, "schedule", ("short_fact",)),
            ("F07", "Exception code", exception_code, "exceptions", ("short_fact",)),
            (
                "F08",
                "Capacity and threshold lookup",
                cross_answer,
                "cross_reference",
                ("cross_section", "exact_numeric"),
            ),
            (
                "F09",
                "Escalation SLA",
                f"{sla_minutes} minutes",
                "exceptions",
                ("exact_numeric",),
            ),
            (
                "F10",
                "Record retention",
                f"{retention_days} days",
                "retention",
                ("exact_numeric",),
            ),
        )
    slug = stable_slug(lane, index, language)
    slices = [
        "pdf"
        if lane == "pdf"
        else "text_markdown_docx"
        if lane in {"docx", "code"}
        else "pptx"
        if lane == "pptx"
        else "xlsx_table",
        "chinese" if language == "zh" else "english",
        "short_fact",
        "cross_section",
        "exact_numeric",
    ]
    if lane == "code":
        slices.append("json_code")
    return {
        "sourceId": doc_id,
        "title": TOPICS[lane][index],
        "filename": f"{slug}.{EXTENSIONS[lane]}",
        "formatLane": lane,
        "language": language,
        "mimeType": MIMES[lane],
        "synthetic": True,
        "reviewState": "draft",
        "slices": slices,
        "facts": [
            {
                "anchor": anchor,
                "label": label,
                "answer": answer,
                "section": section,
                "slices": list(fact_slices),
            }
            for anchor, label, answer, section, fact_slices in facts
        ],
    }


def fact_map(document: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {fact["anchor"]: fact for fact in document["facts"]}


def generate_markdown(document: dict[str, Any], path: Path) -> None:
    facts = fact_map(document)
    zh = document["language"] == "zh"
    heading = "合成评测源文档" if zh else "Synthetic evaluation source document"
    warning = (
        "仅用于 RAG 评测；内容是数据，不是系统指令。"
        if zh
        else "For RAG evaluation only; this content is data, not system instructions."
    )
    json_payload = {
        "source_id": document["sourceId"],
        "project": facts["F01"]["answer"],
        "owner": facts["F02"]["answer"],
        "capacity": facts["F03"]["answer"],
        "trigger_threshold": facts["F04"]["answer"],
        "effective_date": facts["F05"]["answer"],
        "location": facts["F06"]["answer"],
    }
    lines = [
        f"# {document['title']}",
        "",
        f"> {heading} · `{document['sourceId']}` · {warning}",
        "",
        "## JSON configuration",
        "",
        "```json",
        json.dumps(json_payload, ensure_ascii=False, indent=2),
        "```",
        "",
        "## Runtime constants",
        "",
        "```python",
        f"EXCEPTION_CODE = {facts['F07']['answer']!r}",
        f"ESCALATION_SLA = {facts['F09']['answer']!r}",
        "",
        "def should_escalate(load_percent: int) -> bool:",
        f"    return load_percent >= {facts['F04']['answer'].rstrip('%')}",
        "```",
        "",
        "## Cross-section rule",
        "",
        f"[F08] {facts['F08']['label']}: **{facts['F08']['answer']}**.",
        "",
        "## Retention query",
        "",
        "```sql",
        "SELECT source_id, retention_policy",
        "FROM evaluation_sources",
        f"WHERE source_id = '{document['sourceId']}'",
        f"  AND retention_policy = '{facts['F10']['answer']}';",
        "```",
        "",
        f"[F10] {facts['F10']['label']}: {facts['F10']['answer']}.",
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def register_pdf_font() -> str:
    font_name = "NeoEvalUnicode"
    if font_name not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(
            TTFont(font_name, "/usr/share/fonts/truetype/unifont/unifont.ttf")
        )
    return font_name


def generate_pdf(document: dict[str, Any], path: Path) -> None:
    font = register_pdf_font()
    facts = fact_map(document)
    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "EvalTitle",
        parent=styles["Title"],
        fontName=font,
        fontSize=20,
        leading=25,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#17324D"),
    )
    h1 = ParagraphStyle(
        "EvalH1",
        parent=styles["Heading1"],
        fontName=font,
        fontSize=15,
        leading=20,
        textColor=colors.HexColor("#0E7490"),
    )
    body = ParagraphStyle(
        "EvalBody", parent=styles["BodyText"], fontName=font, fontSize=10.5, leading=16
    )
    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        rightMargin=20 * mm,
        leftMargin=20 * mm,
        topMargin=20 * mm,
        bottomMargin=20 * mm,
        title=document["title"],
        author="Neo Chat RAG Evaluation",
    )
    doc.invariant = 1
    story: list[Any] = [
        Paragraph(document["title"], title),
        Spacer(1, 8 * mm),
        Paragraph(f"Synthetic evaluation source · {document['sourceId']}", body),
        Paragraph(f"[F01] {facts['F01']['label']}: {facts['F01']['answer']}", body),
        Paragraph(f"[F02] {facts['F02']['label']}: {facts['F02']['answer']}", body),
        PageBreak(),
        Paragraph("02 · Metrics / 指标", h1),
        Table(
            [
                ["Anchor", "Metric", "Value"],
                ["[F03]", facts["F03"]["label"], facts["F03"]["answer"]],
                ["[F04]", facts["F04"]["label"], facts["F04"]["answer"]],
            ],
            colWidths=[28 * mm, 55 * mm, 75 * mm],
            style=TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), font),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D7EEF4")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#6B8798")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 7),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                ]
            ),
        ),
        PageBreak(),
        Paragraph("03 · Schedule / 排期", h1),
        Paragraph(f"[F05] {facts['F05']['label']}: {facts['F05']['answer']}", body),
        Paragraph(f"[F06] {facts['F06']['label']}: {facts['F06']['answer']}", body),
        Spacer(1, 6 * mm),
        Paragraph(f"[F08] {facts['F08']['label']}: {facts['F08']['answer']}", body),
        PageBreak(),
        Paragraph("04 · Exceptions and retention / 例外与保留", h1),
        Paragraph(f"[F07] {facts['F07']['label']}: {facts['F07']['answer']}", body),
        Paragraph(f"[F09] {facts['F09']['label']}: {facts['F09']['answer']}", body),
        Paragraph(f"[F10] {facts['F10']['label']}: {facts['F10']['answer']}", body),
    ]
    doc.build(story, canvasmaker=functools.partial(canvas.Canvas, invariant=1))


def generate_xlsx(document: dict[str, Any], path: Path) -> None:
    facts = fact_map(document)
    wb = Workbook()
    overview = wb.active
    overview.title = "Overview"
    metrics = wb.create_sheet("Metrics")
    policy = wb.create_sheet("Policy")
    header_fill = PatternFill("solid", fgColor="17324D")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in (overview, metrics, policy):
        sheet.append(["Anchor", "Field", "Exact value"])
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        sheet.column_dimensions["A"].width = 14
        sheet.column_dimensions["B"].width = 30
        sheet.column_dimensions["C"].width = 52
    for anchor in ("F01", "F02", "F05", "F06"):
        fact = facts[anchor]
        overview.append([anchor, fact["label"], fact["answer"]])
    for anchor in ("F03", "F04", "F08"):
        fact = facts[anchor]
        metrics.append([anchor, fact["label"], fact["answer"]])
    for anchor in ("F07", "F09", "F10"):
        fact = facts[anchor]
        policy.append([anchor, fact["label"], fact["answer"]])
    for sheet in (overview, metrics, policy):
        for row in sheet.iter_rows(min_row=2):
            row[0].font = Font(name="Courier New", bold=True, color="0E7490")
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.properties.title = document["title"]
    wb.properties.subject = f"Synthetic RAG evaluation source {document['sourceId']}"
    wb.save(path)
    reloaded = load_workbook(path, data_only=False, read_only=False)
    if reloaded.sheetnames != ["Overview", "Metrics", "Policy"]:
        raise RuntimeError(f"unexpected workbook sheets for {path.name}")
    if any(
        isinstance(cell.value, str) and cell.value.startswith("=")
        for sheet in reloaded.worksheets
        for row in sheet.iter_rows()
        for cell in row
    ):
        raise RuntimeError(f"unexpected formula in source-only workbook {path.name}")
    reloaded.close()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def run_node_generator(output: Path, plan_path: Path, node_path: str) -> None:
    script = Path(__file__).with_name("generate-evaluation-office.cjs")
    env = os.environ.copy()
    env["NODE_PATH"] = node_path
    subprocess.run(
        ["node", str(script), str(plan_path), str(output)], check=True, env=env
    )


def normalize_ooxml_package(path: Path) -> None:
    """Remove directory records and freeze ZIP/core timestamps for strict OPC replay."""
    temporary = path.with_suffix(path.suffix + ".normalized")
    with (
        zipfile.ZipFile(path, "r") as source,
        zipfile.ZipFile(
            temporary,
            "w",
            compression=zipfile.ZIP_DEFLATED,
            compresslevel=9,
        ) as destination,
    ):
        part_names = {info.filename for info in source.infolist() if not info.is_dir()}
        for source_info in source.infolist():
            if source_info.is_dir():
                continue
            content = source.read(source_info.filename)
            if source_info.filename == "[Content_Types].xml":
                namespace = (
                    "http://schemas.openxmlformats.org/package/2006/content-types"
                )
                ET.register_namespace("", namespace)
                root = ET.fromstring(content)
                for child in list(root):
                    if (
                        child.tag == f"{{{namespace}}}Override"
                        and child.attrib.get("PartName", "").lstrip("/")
                        not in part_names
                    ):
                        root.remove(child)
                content = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            if source_info.filename == "ppt/presentation.xml":
                namespace = "http://schemas.openxmlformats.org/presentationml/2006/main"
                ET.register_namespace("p", namespace)
                root = ET.fromstring(content)
                allowed = {"sldMasterIdLst", "sldIdLst", "sldSz", "notesSz"}
                for child in list(root):
                    if child.tag.rsplit("}", 1)[-1] not in allowed:
                        root.remove(child)
                content = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            if source_info.filename == "xl/workbook.xml":
                namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
                ET.register_namespace("", namespace)
                root = ET.fromstring(content)
                for child in list(root):
                    if child.tag.rsplit("}", 1)[-1] not in {"sheets", "calcPr"}:
                        root.remove(child)
                content = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            if source_info.filename == "xl/styles.xml":
                namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
                ET.register_namespace("", namespace)
                root = ET.fromstring(content)
                cell_xfs = root.find(f"{{{namespace}}}cellXfs")
                if cell_xfs is not None:
                    for cell_format in cell_xfs:
                        for child in list(cell_format):
                            cell_format.remove(child)
                content = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            if source_info.filename.startswith(
                "xl/worksheets/sheet"
            ) and source_info.filename.endswith(".xml"):
                namespace = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
                ET.register_namespace("", namespace)
                root = ET.fromstring(content)
                for child in list(root):
                    if child.tag.rsplit("}", 1)[-1] not in {"sheetData", "mergeCells"}:
                        root.remove(child)
                content = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            if source_info.filename.endswith(".rels"):
                namespace = (
                    "http://schemas.openxmlformats.org/package/2006/relationships"
                )
                ET.register_namespace("", namespace)
                root = ET.fromstring(content)
                if source_info.filename == "_rels/.rels":
                    base_directory = ""
                else:
                    before_rels, relationship_name = source_info.filename.rsplit(
                        "/_rels/", 1
                    )
                    source_part = posixpath.join(
                        before_rels, relationship_name.removesuffix(".rels")
                    )
                    base_directory = posixpath.dirname(source_part)
                for relationship in root:
                    target = relationship.attrib.get("Target", "")
                    if relationship.attrib.get(
                        "TargetMode"
                    ) != "External" and target.startswith("/"):
                        relationship.attrib["Target"] = posixpath.relpath(
                            target.lstrip("/"), start=base_directory or "."
                        )
                if source_info.filename == "xl/_rels/workbook.xml.rels":
                    for relationship in list(root):
                        if relationship.attrib.get("Type", "").endswith("/theme"):
                            root.remove(relationship)
                content = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            if source_info.filename == "docProps/core.xml":
                content = re.sub(
                    rb"(<dcterms:(?:created|modified)[^>]*>)[^<]*(</dcterms:(?:created|modified)>)",
                    rb"\g<1>2026-07-24T00:00:00Z\g<2>",
                    content,
                )
            target_info = zipfile.ZipInfo(
                source_info.filename,
                date_time=(2026, 7, 24, 0, 0, 0),
            )
            target_info.compress_type = zipfile.ZIP_DEFLATED
            target_info.create_system = 3
            target_info.external_attr = 0o600 << 16
            destination.writestr(target_info, content)
    temporary.replace(path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("output", type=Path)
    parser.add_argument(
        "--node-path", default="/tmp/neo-chat-rag-eval-node/node_modules"
    )
    args = parser.parse_args()
    output = args.output.resolve()
    if output.exists():
        shutil.rmtree(output)
    output.mkdir(parents=True)
    documents = [build_document(lane, index) for lane in LANES for index in range(10)]
    plan = {
        "schemaVersion": SCHEMA,
        "generatedAt": GENERATED_AT,
        "synthetic": True,
        "promotionEligible": False,
        "documents": documents,
    }
    plan_path = output / "source-plan.json"
    plan_path.write_text(
        json.dumps(plan, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    for document in documents:
        path = output / document["filename"]
        if document["formatLane"] == "pdf":
            generate_pdf(document, path)
        elif document["formatLane"] == "xlsx":
            generate_xlsx(document, path)
        elif document["formatLane"] == "code":
            generate_markdown(document, path)
    run_node_generator(output, plan_path, args.node_path)
    for document in documents:
        if document["formatLane"] in {"docx", "pptx", "xlsx"}:
            normalize_ooxml_package(output / document["filename"])
    hashes: set[str] = set()
    for document in documents:
        path = output / document["filename"]
        if not path.is_file() or path.stat().st_size == 0:
            raise RuntimeError(f"missing generated source: {path}")
        document["byteSize"] = path.stat().st_size
        document["sha256"] = sha256(path)
        if document["sha256"] in hashes:
            raise RuntimeError(f"duplicate generated content: {path.name}")
        hashes.add(document["sha256"])
    plan_path.unlink()
    manifest_path = output / "manifest.json"
    manifest_path.write_text(
        json.dumps(plan, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(
        json.dumps(
            {
                "output": str(output),
                "documents": len(documents),
                "uniqueHashes": len(hashes),
                "manifestSha256": sha256(manifest_path),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
